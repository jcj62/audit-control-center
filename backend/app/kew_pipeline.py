from __future__ import annotations

import io
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


def parse_value(value: object) -> tuple[float, str]:
    try:
        raw = str(value).strip()
        lowered = raw.lower().replace("<", "")
        if "k" in lowered:
            numeric = float(lowered.replace("k", "")) * 1000
        else:
            numeric = float(lowered)
        return numeric, raw
    except Exception:
        return 0, str(value)


def extract_rating(text: object) -> str:
    match = re.search(r"\d+\s*A", str(text))
    return match.group() if match else ""


def generate_remark(entry: dict, table_type: str) -> str:
    try:
        if table_type == "VOLT":
            l, _ = parse_value(entry.get("L-N Voltage[V]"))
            ne, _ = parse_value(entry.get("N-PE Voltage[V]"))

            if ne >= 180:
                if l <= 3:
                    return "L/N reversed (Line low, N-E high)"
                return "Possible L/N reversal"

            if 10 <= ne <= 70:
                return "Earthing issue (N-E high)"

            if l < 200 or l > 250:
                return "Voltage out of range"

            return "Within Limits"

        if table_type == "LOOP":
            z, _ = parse_value(entry.get("EFLI[Ω]"))
            pfc, _ = parse_value(entry.get("PFC[A]"))
            breaker = str(entry.get("Breaker Rating", "")).upper()

            if z == 0:
                return "Invalid loop reading"

            nums = re.findall(r"\d+", breaker)
            current_rating = int(nums[0]) if nums else None

            multiplier = None
            if breaker.startswith("B"):
                multiplier = 5
            elif breaker.startswith("C"):
                multiplier = 10
            elif breaker.startswith("D"):
                multiplier = 20

            if multiplier and current_rating and pfc:
                required = current_rating * multiplier
                if pfc < required:
                    return f"Loop high (PFC {int(pfc)}A < {required}A)"
                return "OK"

            if current_rating and pfc and pfc < current_rating:
                return f"Low fault level (PFC {int(pfc)}A < {current_rating}A)"

            return "OK"

        if table_type == "RCD":
            time = entry.get("Tripping Time", "")
            current = entry.get("Tripping Current", "")
            rating = str(entry.get("RCCB/ RCBO Rating", "")).lower()
            asset = str(entry.get("DB/ Socket", "")).lower()

            try:
                trip_time = float(str(time).replace("ms", "").strip())
            except Exception:
                trip_time = None

            try:
                trip_current = float(str(current).replace("mA", "").strip())
            except Exception:
                trip_current = None

            nums = re.findall(r"\d+", rating)
            rated = int(nums[0]) if nums else None

            if rated == 100:
                if "acdb" not in asset and "pdb" not in asset:
                    return "Working OK. 100mA RCCB to be replaced with 30mA RCBO"
                return "OK"

            if rated == 30 and trip_current is not None and trip_current <= 21:
                leakage = rated - trip_current
                return f"Working OK. Suspected {int(leakage)}mA leakage, to be checked"

            if trip_time is not None and trip_time > 300:
                return "RCD trip delay high"

            return "OK"

        if table_type == "INSU":
            value = entry.get("Insulation Resistance", "")
            try:
                resistance = float(str(value).replace("MΩ", "").strip())
                if resistance < 1:
                    return "Low insulation resistance"
            except Exception:
                pass
            return "OK"

        if table_type == "CONST":
            value = entry.get("Continuity Resistance", "")
            try:
                resistance = float(str(value).replace("Ω", "").strip())
                if resistance > 1:
                    return "High continuity resistance"
            except Exception:
                pass
            return "OK"

    except Exception:
        return "Check Data"

    return "OK"


def _decode_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    return text.splitlines()


def parse_kew_content(content: bytes) -> tuple[list, list, list, list, list, list, list]:
    panel_voltage: list[dict] = []
    panel_loop: list[dict] = []
    db_voltage: list[dict] = []
    db_loop: list[dict] = []
    rcd_table: list[dict] = []
    insu_table: list[dict] = []
    const_table: list[dict] = []

    headers: list[str] = []
    last_rcd: dict | None = None

    for line in _decode_csv(content):
        line = line.strip()

        if line.startswith("No") and "Function" in line:
            headers = [header.strip() for header in line.split(",")]
            continue

        if not line or line.startswith("No"):
            continue

        values = [value.strip() for value in line.split(",")]
        if len(values) < 5 or not headers:
            continue

        data = dict(zip(headers, values))
        function = str(data.get("Function", "")).upper()
        location = data.get("Comment1", "")
        asset = data.get("Comment2", "")

        if "VOLT" in function:
            l, l_display = parse_value(data.get("L-N[V]"))
            ne, ne_display = parse_value(data.get("N-PE[V]"))
            _, lpe_display = parse_value(data.get("L-PE[V]"))

            entry = {
                "Sl. No.": data.get("No"),
                "Location": location,
                "L-PE Voltage[V]": lpe_display.replace("+", "").strip(),
                "L-N Voltage[V]": l_display.replace("+", "").strip(),
                "N-PE Voltage[V]": ne_display.replace("+", "").strip(),
                "Remarks": generate_remark(
                    {
                        "L-N Voltage[V]": l,
                        "N-PE Voltage[V]": ne,
                    },
                    "VOLT",
                ),
            }

            if "panel" in asset.lower():
                entry["Panel/ Feeder"] = asset
                panel_voltage.append(entry)
            else:
                entry["DB/ Socket"] = asset
                db_voltage.append(entry)

        elif "LOOP" in function:
            _, z_display = parse_value(data.get("LOOP[Ω]"))
            _, pfc_display = parse_value(data.get("PFC[A]"))
            _, psc_display = parse_value(data.get("PSC[A]"))
            _, ln_display = parse_value(data.get("L-N[Ω]"))
            _, mains_display = parse_value(data.get("Mains[V]"))

            entry = {
                "Sl. No.": data.get("No"),
                "Location": location,
                "Breaker Rating": extract_rating(asset),
                "EFLI[Ω]": z_display,
                "PFC[A]": pfc_display,
                "PSC[A]": psc_display,
                "L-N[Ω]": ln_display,
                "Mains Voltage[V]": mains_display,
            }
            entry["Remarks"] = generate_remark(entry, "LOOP")

            if "panel" in asset.lower():
                entry["Panel/ Feeder"] = asset
                panel_loop.append(entry)
            else:
                entry["DB/ Socket"] = asset
                db_loop.append(entry)

        elif "INSU" in function:
            insu_table.append(
                {
                    "Sl. No.": data.get("No"),
                    "Location": location if location else "Ground Floor, Maintenance Room",
                    "Asset": asset,
                    "Insulation Resistance": str(data.get("INSU[MΩ]", "")).strip(),
                    "Test Voltage (V)": data.get("Range") or data.get("Out[V]"),
                    "Remarks": data.get("PAT") or "OK",
                }
            )

        elif "CONST" in function:
            raw_value = str(data.get("CONST", "")).strip()
            if not raw_value:
                for value in values:
                    if "Ω" in value:
                        raw_value = value
                        break

            nums = re.findall(r"\d+\.?\d*", raw_value)
            parsed = float(nums[0]) if nums else None

            entry = {
                "Sl. No.": data.get("No"),
                "Location": location,
                "Asset": asset,
                "Continuity Resistance": f"{parsed} Ω" if parsed is not None else "",
            }
            entry["Remarks"] = generate_remark(entry, "CONST")
            const_table.append(entry)

        elif "RCD" in function:
            raw_value = str(data.get("RCD", "")).strip()
            nums = re.findall(r"\d+\.?\d*", raw_value)
            parsed = float(nums[0]) if nums else None

            if "RAMP" in function:
                last_rcd = {
                    "Sl. No.": data.get("No"),
                    "Location": location,
                    "DB/ Socket": asset,
                    "Tripping Current": f"{parsed} mA" if parsed is not None else "",
                    "Tripping Time": "",
                    "RCCB/ RCBO Rating": data.get("Idn"),
                    "Remarks": "OK",
                }
            elif last_rcd:
                last_rcd["Tripping Time"] = f"{parsed} ms" if parsed is not None else ""
                if not last_rcd["Location"]:
                    last_rcd["Location"] = location
                if not last_rcd["DB/ Socket"]:
                    last_rcd["DB/ Socket"] = asset
                last_rcd["Remarks"] = generate_remark(last_rcd, "RCD")
                rcd_table.append(last_rcd)
                last_rcd = None

    return panel_voltage, panel_loop, db_voltage, db_loop, rcd_table, insu_table, const_table


def build_kew_workbook(files: list[tuple[str, bytes]], output_dir: Path, output_name: str) -> Path:
    all_panel_voltage: list[dict] = []
    all_panel_loop: list[dict] = []
    all_db_voltage: list[dict] = []
    all_db_loop: list[dict] = []
    all_rcd: list[dict] = []
    all_insu: list[dict] = []
    all_const: list[dict] = []

    for _, content in files:
        pv, pl, dv, dl, rcd, insu, const = parse_kew_content(content)
        all_panel_voltage.extend(pv)
        all_panel_loop.extend(pl)
        all_db_voltage.extend(dv)
        all_db_loop.extend(dl)
        all_rcd.extend(rcd)
        all_insu.extend(insu)
        all_const.extend(const)

    workbook = Workbook()
    workbook.remove(workbook.active)

    tables = {
        "Panel Voltage": pd.DataFrame(all_panel_voltage),
        "Panel Loop": pd.DataFrame(all_panel_loop),
        "DB Voltage": pd.DataFrame(all_db_voltage),
        "DB Loop": pd.DataFrame(all_db_loop),
        "RCD": pd.DataFrame(all_rcd),
        "INSUL": pd.DataFrame(all_insu),
        "CONST": pd.DataFrame(all_const),
    }

    column_orders = {
        "Panel Voltage": ["Sl. No.", "Location", "Panel/ Feeder", "L-PE Voltage[V]", "L-N Voltage[V]", "N-PE Voltage[V]", "Remarks"],
        "DB Voltage": ["Sl. No.", "Location", "DB/ Socket", "L-PE Voltage[V]", "L-N Voltage[V]", "N-PE Voltage[V]", "Remarks"],
        "Panel Loop": ["Sl. No.", "Location", "Panel/ Feeder", "Breaker Rating", "EFLI[Ω]", "PFC[A]", "PSC[A]", "L-N[Ω]", "Mains Voltage[V]", "Remarks"],
        "DB Loop": ["Sl. No.", "Location", "DB/ Socket", "Breaker Rating", "EFLI[Ω]", "PFC[A]", "PSC[A]", "L-N[Ω]", "Mains Voltage[V]", "Remarks"],
        "RCD": ["Sl. No.", "Location", "DB/ Socket", "Tripping Current", "Tripping Time", "RCCB/ RCBO Rating", "Remarks"],
        "INSUL": ["Sl. No.", "Location", "Asset", "Insulation Resistance", "Test Voltage (V)", "Remarks"],
        "CONST": ["Sl. No.", "Location", "Asset", "Continuity Resistance", "Remarks"],
    }

    red = Font(color="FF0000")

    for name, dataframe in tables.items():
        worksheet = workbook.create_sheet(title=name)
        if dataframe.empty:
            continue

        dataframe = dataframe.reindex(columns=column_orders[name])
        worksheet.append(list(dataframe.columns))

        for row_data in dataframe.itertuples(index=False):
            worksheet.append(list(row_data))

        headers = list(dataframe.columns)
        for row in worksheet.iter_rows(min_row=2):
            remark = row[headers.index("Remarks")].value
            if remark in ["OK", "Within Limits", None]:
                continue

            remark = str(remark)
            if "Voltage" in name:
                if "Voltage out of range" in remark:
                    row[headers.index("L-N Voltage[V]")].font = red
                elif "Earthing issue" in remark or "N-E" in remark or "reversed" in remark:
                    row[headers.index("N-PE Voltage[V]")].font = red
            elif "Loop" in name:
                if "Loop high" in remark or "Low fault level" in remark:
                    row[headers.index("EFLI[Ω]")].font = red
            elif name == "RCD":
                if "leakage" in remark:
                    row[headers.index("Tripping Current")].font = red
                elif "delay" in remark:
                    row[headers.index("Tripping Time")].font = red
                elif "100mA" in remark:
                    row[headers.index("RCCB/ RCBO Rating")].font = red
            elif name == "INSUL" and "Low insulation" in remark:
                row[headers.index("Insulation Resistance")].font = red
            elif name == "CONST" and "High continuity" in remark:
                row[headers.index("Continuity Resistance")].font = red

    safe_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", (output_name or "kew_report").strip()).strip("_") or "kew_report"
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"{safe_name}.xlsx"
    workbook.save(file_path)
    return file_path
